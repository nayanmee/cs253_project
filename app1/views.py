from django.shortcuts import render,redirect
from app1.models import Profile,Student,Pending_Requests
from django.contrib.auth.models import User, auth
#import . from models
import openpyxl


def upload(request):
    if "GET" == request.method:
        return render(request, 'upload.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)

        worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()

        for row in worksheet.iter_rows():
            print(int(row[0].value))
            print(str(row[1].value))
            p=Student.objects.create(number   =  int(row[0].value))

            p.first_name = str(row[1].value)
            p.last_name = str(row[2].value)
            p.department = str(row[3].value)
            p.save()
            print(p.number) 


        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        return render(request, 'upload.html', {"excel_data":excel_data})



def register(request):
    if request.method == 'POST':
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        username = request.POST['username']
        email = request.POST['email']
        role=request.POST['role']
        password1 = request.POST['password1']
        password2 = request.POST['password2']

        Admin = User.objects.create_user(username='Admin',password='a',email='admin@gmail.com',first_name='Admin',last_name='m')
        Admin.save()
        Pending_Requests.objects.create(first_name = first_name,last_name=last_name,username=username,email=email,password = password1,role =role)
        Admin_profile=Profile.objects.create(user = Admin,role='Admin')
        
        # pro.start_No=41
        # pro.end_No=44

      
        # pro.save()
        # print("hi")

        # print(pro.role)
        # print(pro.user.username)

        #add messages latter
        return redirect('/login')
    else:
        return render(request,'register.html')


def login(request):
    if(request.method=="POST"):
        print("check_2")
        username=request.POST['username']
        password=request.POST['password']

        user=auth.authenticate(username=username,password=password)
        request_ta = list()
        request_manager = list()
        for r in Pending_Requests.objects.all():
            if(r.role == 'Ta'):
                stud1 = r
                request_ta.append(stud1)
            if(r.role == "Manager"):
                stud2 = r
                request_manager.append(stud2)    
        
                

        if user is not None:
            auth.login(request,user)
            actual_user = Profile.objects.get(user=user)
            i=0
            print (actual_user.role)
            if(actual_user.role=='Ta'):
                i=i+1
                return render(request,'ta_dashboard.html',{'user':actual_user},)

            #     #code for ta
            #     arr = list()
            #     for num in range(actual_user.start_No,actual_user.end_No+1):

            #         stu = Student.objects.get(number=num)
            #         print(stu.first_name)
            #         arr.append(stu) 

            # #print(arr[0].first_name)

            #     return render(request,'ta_dashboard.html',{'user':actual_user,"arr":arr})
            
            elif(actual_user.role=='Manager'):
                i=i+1
                #code for manager
                return render(request ,'manager_dashboard.html',{ "request_ta":request_ta })
            else:
                i=i+1
                #code for admin
                return render(request ,'admin_dashboard.html',{ "request_manager":request_manager })

           
            #return render(request ,'manager_dashboard.html',{ "request_ta":request_ta })
            
        else:
            print('invalid Credentials')
            return render(request,'login.html')
    else:
        print("check4")
        return render(request,'login.html')


def student(request,student_id):
    if(request.method=="GET"):
        user=auth.get_user(request)
        if(user.is_authenticated):
            actual_user = Profile.objects.get(user=user)
            if(actual_user.start_No<=student_id or actual_user.end_No>=student_id):
                stu=Student.objects.get(number=student_id)
                return render(request,'student.html',{'user':actual_user,'student':stu})
            else:
                return redirect('/login')
                #this redirection supposed to be changed 

        #    return redirect('/login')
        #print("jii")
        else:
            return redirect('/login')


    else:
        discrepency=request.POST.get('discrepency')
        
        comment=request.POST['comment']
        stu=Student.objects.get(number=student_id)
        stu.check_box='Yes'
        print(stu.discrepency)
        if(str(discrepency)=='on'):
            stu.discrepency='Yes'
        else:
            stu.discrepency='No'


        print(stu.discrepency)
        stu.comment=comment
        stu.save()
        user=auth.get_user(request)
        actual_user = Profile.objects.get(user=user)
        arr=list()
        for num in range(actual_user.start_No,actual_user.end_No+1):
            student = Student.objects.get(number=num)
            print(student.first_name)
            arr.append(student)

        return render(request,'ta_dashboard.html',{'user':actual_user,"arr":arr})



    








def logout(request):
    print(auth.get_user(request).username)

    print("logout working")
    auth.logout(request)
    print(auth.get_user(request).username)
    if(auth.get_user(request).is_authenticated):
        print("hii")
    else:
        print("bye")
    return redirect('/login')

def Pending_Request(request):
        if request.method == 'POST':
            for r in Pending_Requests.objects.all():
                r.status = request.POST.get(r.username)

                if (r.status == 'ACCEPT'):
                    user_dummy = User.objects.create_user(username=r.username,password=r.password,email=r.email,first_name=r.first_name,last_name=r.last_name)
                    user_dummy.save()
                    pro = Profile(user = user_dummy,role = r.role)
                    pro.save()
                    r.delete()
                if (r.status == 'REJECT'):
                    
                    r.delete()


             
        request_ta = list()
        request_manager = list()             

        for r in Pending_Requests.objects.all():
                if(r.role == 'Ta'):
                    stud1 = r
                    request_ta.append(stud1)
                if(r.role == "Manager"):
                    stud2 = r
                    request_manager.append(stud2) 
        
        ROLE = request.POST.get('ROLE', None)
                      
        if(ROLE == 'Admin'):
            return render(request ,'admin_dashboard.html',{ "request_manager":request_manager })
        else:
            return render(request ,'manager_dashboard.html',{ "request_ta":request_ta})
         
         
       


                
                

